"""Excel dosya okuyucu modülü."""

from pathlib import Path
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None


@dataclass
class ExcelSheet:
    """Excel sayfası içeriği."""
    name: str
    data: List[List[Any]]
    headers: List[str]
    row_count: int
    column_count: int
    has_merged_cells: bool = False


@dataclass
class ExcelContent:
    """Excel dosyası içeriği."""
    source: str
    filename: str
    sheet_count: int
    sheets: List[ExcelSheet] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelParser:
    """Excel dosya okuyucu sınıfı."""

    def __init__(self):
        self._current_wb = None
        if pd is None:
            raise ImportError("pandas kütüphanesi yüklü değil. 'pip install pandas openpyxl' komutunu çalıştırın.")

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - acik workbook'lari kapat."""
        self.close()
        return False

    def close(self):
        """Acik kaynaklari temizle."""
        if self._current_wb is not None:
            try:
                self._current_wb.close()
            except (AttributeError, TypeError):
                pass
            self._current_wb = None

    def parse(self, file_path: str) -> ExcelContent:
        """Excel dosyasını oku ve içeriği çıkar."""
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Dosya bulunamadı: {file_path}")

        extension = file_path.suffix.lower()

        content = ExcelContent(
            source=str(file_path),
            filename=file_path.name,
            sheet_count=0
        )

        if extension == '.csv':
            # CSV dosyası
            content = self._parse_csv(file_path, content)
        else:
            # Excel dosyası (.xlsx, .xls)
            content = self._parse_excel(file_path, content)

        return content

    def _parse_csv(self, file_path: Path, content: ExcelContent) -> ExcelContent:
        """CSV dosyasını oku."""
        try:
            # Encoding'i otomatik tespit etmeye çalış
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-9', 'cp1254']

            df = None
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if df is None:
                df = pd.read_csv(file_path, encoding='utf-8', errors='replace')

            sheet = ExcelSheet(
                name="Sheet1",
                headers=list(df.columns),
                data=df.values.tolist(),
                row_count=len(df),
                column_count=len(df.columns)
            )

            content.sheets.append(sheet)
            content.sheet_count = 1

        except Exception as e:
            content.metadata['error'] = str(e)

        return content

    def _parse_excel(self, file_path: Path, content: ExcelContent) -> ExcelContent:
        """Excel dosyasını oku."""
        try:
            # Tüm sayfaları oku
            excel_file = pd.ExcelFile(file_path)
            content.sheet_count = len(excel_file.sheet_names)

            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)

                    # NaN değerleri boş stringe çevir
                    df = df.fillna("")

                    sheet = ExcelSheet(
                        name=sheet_name,
                        headers=[str(col) for col in df.columns],
                        data=[[str(cell) if cell != "" else "" for cell in row] for row in df.values.tolist()],
                        row_count=len(df),
                        column_count=len(df.columns)
                    )

                    # Birleşik hücre kontrolü (openpyxl ile)
                    if load_workbook is not None and file_path.suffix.lower() == '.xlsx':
                        try:
                            wb = load_workbook(file_path, data_only=True)
                            ws = wb[sheet_name]
                            sheet.has_merged_cells = len(ws.merged_cells.ranges) > 0
                            wb.close()
                        except (KeyError, IOError, ValueError) as e:
                            # Workbook/sayfa okuma hatasi
                            pass

                    content.sheets.append(sheet)

                except (ValueError, TypeError, KeyError) as e:
                    # Sayfa okunamazsa atla
                    continue

            # Metadata
            if load_workbook is not None and file_path.suffix.lower() == '.xlsx':
                try:
                    wb = load_workbook(file_path, data_only=True)
                    props = wb.properties
                    if props:
                        content.metadata = {
                            'title': props.title or "",
                            'creator': props.creator or "",
                            'created': str(props.created) if props.created else "",
                            'modified': str(props.modified) if props.modified else ""
                        }
                    wb.close()
                except (IOError, KeyError, AttributeError) as e:
                    # Metadata okuma hatasi
                    pass

        except Exception as e:
            content.metadata['error'] = str(e)

        return content

    def get_sheet_as_dataframe(self, file_path: str, sheet_name: str = None) -> Optional[Any]:
        """Belirli bir sayfayı DataFrame olarak al."""
        try:
            if sheet_name:
                return pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                return pd.read_excel(file_path)
        except (IOError, ValueError, KeyError) as e:
            # Excel okuma hatasi
            return None

    def get_summary(self, content: ExcelContent) -> Dict[str, Any]:
        """Excel içeriğinin özetini al."""
        total_rows = sum(sheet.row_count for sheet in content.sheets)
        total_columns = max((sheet.column_count for sheet in content.sheets), default=0)

        return {
            "filename": content.filename,
            "sheet_count": content.sheet_count,
            "total_rows": total_rows,
            "max_columns": total_columns,
            "sheet_names": [sheet.name for sheet in content.sheets]
        }

    def to_dict(self, content: ExcelContent) -> Dict[str, Any]:
        """ExcelContent'i sözlüğe çevir."""
        return {
            "type": "excel",
            "source": content.source,
            "filename": content.filename,
            "sheet_count": content.sheet_count,
            "metadata": content.metadata,
            "sheets": [
                {
                    "name": sheet.name,
                    "headers": sheet.headers,
                    "data": sheet.data,
                    "row_count": sheet.row_count,
                    "column_count": sheet.column_count,
                    "has_merged_cells": sheet.has_merged_cells
                }
                for sheet in content.sheets
            ]
        }

    def to_markdown_tables(self, content: ExcelContent) -> str:
        """Excel içeriğini Markdown tablo formatına çevir."""
        markdown = []

        for sheet in content.sheets:
            markdown.append(f"### {sheet.name}\n")

            if not sheet.headers:
                continue

            # Başlık satırı
            header_row = "| " + " | ".join(sheet.headers) + " |"
            separator = "| " + " | ".join(["---"] * len(sheet.headers)) + " |"

            markdown.append(header_row)
            markdown.append(separator)

            # Veri satırları (ilk 50 satır)
            for row in sheet.data[:50]:
                row_str = "| " + " | ".join(str(cell)[:50] for cell in row) + " |"
                markdown.append(row_str)

            if len(sheet.data) > 50:
                markdown.append(f"\n*... ve {len(sheet.data) - 50} satır daha*\n")

            markdown.append("")

        return "\n".join(markdown)
